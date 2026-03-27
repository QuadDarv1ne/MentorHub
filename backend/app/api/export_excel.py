"""
Excel Export for User Data

Converts user data to Excel format using openpyxl.
"""

import io
from datetime import datetime
from typing import Dict, Any, List

from fastapi import HTTPException, Response


def export_as_excel(data: Dict[str, Any]) -> Response:
    """
    Export user data to Excel format.
    
    Args:
        data: User data dictionary from collect_user_data()
    
    Returns:
        Response with Excel file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=503, detail="Excel export not available (openpyxl not installed)")
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left_side=Side(style='thin'),
        right_side=Side(style='thin'),
        top_side=Side(style='thin'),
        bottom_side=Side(style='thin')
    )
    
    # Sheet 1: User Profile
    ws_user = wb.active
    ws_user.title = "User Profile"
    
    user_data_rows = [
        ["Field", "Value"],
        ["Username", data["user"]["username"]],
        ["Email", data["user"]["email"]],
        ["Full Name", data["user"]["full_name"] or "N/A"],
        ["Role", data["user"]["role"]],
        ["Is Active", str(data["user"]["is_active"])],
        ["Is Verified", str(data["user"]["is_verified"])],
        ["Export Date", data["export_date"]]
    ]
    
    _write_excel_sheet(ws_user, user_data_rows, header_font, header_fill, header_alignment, thin_border)
    ws_user.column_dimensions['A'].width = 20
    ws_user.column_dimensions['B'].width = 40
    
    # Sheet 2: Sessions
    ws_sessions = wb.create_sheet(title="Sessions")
    sessions_data = [["ID", "Student ID", "Mentor ID", "Scheduled At", "Duration", "Status"]]
    for session in data["sessions"]:
        sessions_data.append([
            session["id"],
            session["student_id"],
            session["mentor_id"],
            session["scheduled_at"],
            session["duration_minutes"],
            session["status"]
        ])
    _write_excel_sheet(ws_sessions, sessions_data, header_font, header_fill, header_alignment, thin_border)
    
    # Sheet 3: Payments
    ws_payments = wb.create_sheet(title="Payments")
    payments_data = [["ID", "Student ID", "Mentor ID", "Amount", "Currency", "Status", "Created At"]]
    for payment in data["payments"]:
        payments_data.append([
            payment["id"],
            payment["student_id"],
            payment["mentor_id"],
            payment["amount"],
            payment["currency"],
            payment["status"],
            payment["created_at"]
        ])
    _write_excel_sheet(ws_payments, payments_data, header_font, header_fill, header_alignment, thin_border)
    
    # Sheet 4: Statistics
    ws_stats = wb.create_sheet(title="Statistics")
    stats_data = [
        ["Category", "Count"],
        ["Sessions", len(data["sessions"])],
        ["Payments", len(data["payments"])],
        ["Reviews", len(data["reviews"])],
        ["Progress Records", len(data["progress"])],
        ["Achievements", len(data["achievements"])],
        ["Messages", len(data["messages"])],
        ["Course Enrollments", len(data["enrollments"])],
        ["Total", sum([
            len(data["sessions"]),
            len(data["payments"]),
            len(data["reviews"]),
            len(data["progress"]),
            len(data["achievements"]),
            len(data["messages"]),
            len(data["enrollments"])
        ])]
    ]
    _write_excel_sheet(ws_stats, stats_data, header_font, header_fill, header_alignment, thin_border)
    
    # Save
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"mentorhub_export_{data['user']['username']}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def _write_excel_sheet(ws, data: List[List], header_font, header_fill, header_alignment, thin_border):
    """
    Helper function to write Excel sheet with styling.
    
    Args:
        ws: Worksheet object
        data: 2D list of data to write
        header_font: Font style for header
        header_fill: Fill style for header
        header_alignment: Alignment style for header
        thin_border: Border style for cells
    """
    from openpyxl.utils import get_column_letter
    
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            else:
                cell.border = thin_border
    
    # Adjust column widths
    for col_idx in range(1, len(data[0]) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
