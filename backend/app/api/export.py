"""
User Data Export API
Экспорт данных пользователя в формате JSON/CSV/PDF/Excel (GDPR compliance)
"""

import json
import csv
import io
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse, JSONResponse, Response
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import select

from app.dependencies import get_db, get_current_user
from app.models.user import User
from app.models.session import Session as SessionModel
from app.models.payment import Payment
from app.models.review import Review
from app.models.progress import Progress
from app.models.achievement import Achievement
from app.models.message import Message
from app.models.course import Course, CourseEnrollment
from app.utils.cache import cached

router = APIRouter(prefix="/export", tags=["Data Export"])


try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@router.get("/data", summary="Экспорт всех данных пользователя")
async def export_user_data(
    format: str = "json",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Экспорт всех данных пользователя в формате JSON, CSV, PDF или Excel.

    Включает:
    - Профиль пользователя
    - История сессий
    - История платежей
    - Отзывы
    - Прогресс обучения
    - Достижения
    - Сообщения
    - Записки на курсы

    GDPR compliance: пользователь может запросить копию своих данных.
    """
    
    # Получаем все данные пользователя
    user_data = {
        "export_date": datetime.utcnow().isoformat(),
        "user": {
            "id": current_user.id,
            "username": current_user.username,
            "email": current_user.email,
            "full_name": current_user.full_name,
            "role": current_user.role.value,
            "is_active": current_user.is_active,
            "is_verified": current_user.is_verified,
            "created_at": current_user.created_at.isoformat() if current_user.created_at else None,
            "updated_at": current_user.updated_at.isoformat() if current_user.updated_at else None,
            "bio": current_user.bio,
            "phone": current_user.phone,
            "timezone": current_user.timezone,
            "language": current_user.language,
        },
        "sessions": [],
        "payments": [],
        "reviews": [],
        "progress": [],
        "achievements": [],
        "messages": [],
        "enrollments": [],
    }
    
    # Сессии (как студент и как ментор)
    sessions_query = select(SessionModel).where(
        (SessionModel.student_id == current_user.id) | 
        (SessionModel.mentor_id == current_user.id)
    )
    sessions = db.execute(sessions_query).scalars().all()
    
    for session in sessions:
        user_data["sessions"].append({
            "id": session.id,
            "student_id": session.student_id,
            "mentor_id": session.mentor_id,
            "scheduled_at": session.scheduled_at.isoformat() if session.scheduled_at else None,
            "duration_minutes": session.duration_minutes,
            "status": session.status.value,
            "meeting_link": session.meeting_link,
            "notes": session.notes,
            "created_at": session.created_at.isoformat() if session.created_at else None,
        })
    
    # Платежи
    payments_query = select(Payment).where(
        (Payment.student_id == current_user.id) | 
        (Payment.mentor_id == current_user.id)
    )
    payments = db.execute(payments_query).scalars().all()
    
    for payment in payments:
        user_data["payments"].append({
            "id": payment.id,
            "student_id": payment.student_id,
            "mentor_id": payment.mentor_id,
            "session_id": payment.session_id,
            "amount": float(payment.amount),
            "currency": payment.currency,
            "status": payment.status.value,
            "payment_method": payment.payment_method,
            "created_at": payment.created_at.isoformat() if payment.created_at else None,
        })
    
    # Отзывы
    reviews_query = select(Review).where(Review.user_id == current_user.id)
    reviews = db.execute(reviews_query).scalars().all()
    
    for review in reviews:
        user_data["reviews"].append({
            "id": review.id,
            "user_id": review.user_id,
            "course_id": review.course_id,
            "rating": review.rating,
            "comment": review.comment,
            "created_at": review.created_at.isoformat() if review.created_at else None,
        })
    
    # Прогресс
    progress_query = select(Progress).where(Progress.user_id == current_user.id)
    progress = db.execute(progress_query).scalars().all()
    
    for prog in progress:
        user_data["progress"].append({
            "id": prog.id,
            "user_id": prog.user_id,
            "course_id": prog.course_id,
            "lesson_id": prog.lesson_id,
            "progress_percent": prog.progress_percent,
            "completed": prog.completed,
            "created_at": prog.created_at.isoformat() if prog.created_at else None,
        })
    
    # Достижения
    achievements_query = select(Achievement).where(Achievement.user_id == current_user.id)
    achievements = db.execute(achievements_query).scalars().all()
    
    for achievement in achievements:
        user_data["achievements"].append({
            "id": achievement.id,
            "user_id": achievement.user_id,
            "achievement_id": achievement.achievement_id,
            "earned_at": achievement.earned_at.isoformat() if achievement.earned_at else None,
        })
    
    # Сообщения (отправленные и полученные)
    messages_query = select(Message).where(
        (Message.sender_id == current_user.id) | 
        (Message.recipient_id == current_user.id)
    )
    messages = db.execute(messages_query).scalars().limit(100).all()  # Лимит 100 сообщений
    
    for message in messages:
        user_data["messages"].append({
            "id": message.id,
            "sender_id": message.sender_id,
            "recipient_id": message.recipient_id,
            "content": message.content[:100] + "..." if len(message.content) > 100 else message.content,  # Обрезаем длинные сообщения
            "is_read": message.is_read,
            "created_at": message.created_at.isoformat() if message.created_at else None,
        })
    
    # Записки на курсы
    enrollments_query = select(CourseEnrollment).where(CourseEnrollment.user_id == current_user.id)
    enrollments = db.execute(enrollments_query).scalars().all()
    
    for enrollment in enrollments:
        user_data["enrollments"].append({
            "id": enrollment.id,
            "user_id": enrollment.user_id,
            "course_id": enrollment.course_id,
            "enrolled_at": enrollment.enrolled_at.isoformat() if enrollment.enrolled_at else None,
            "completed": enrollment.completed,
            "completed_at": enrollment.completed_at.isoformat() if enrollment.completed_at else None,
        })
    
    # Возвращаем данные в нужном формате
    format_lower = format.lower()
    
    if format_lower == "csv":
        return _export_as_csv(user_data)
    elif format_lower == "pdf":
        return _export_as_pdf(user_data)
    elif format_lower == "excel" or format_lower == "xlsx":
        return _export_as_excel(user_data)
    else:
        return JSONResponse(
            content=user_data,
            headers={
                "Content-Disposition": f'attachment; filename="mentorhub_data_export_{current_user.id}_{datetime.utcnow().strftime("%Y%m%d")}.json"'
            }
        )


def _export_as_csv(data: dict) -> StreamingResponse:
    """Конвертация данных в CSV формат"""
    output = io.StringIO()
    
    # Экспортируем основную информацию о пользователе
    writer = csv.writer(output)
    writer.writerow(["Field", "Value"])
    writer.writerow(["Username", data["user"]["username"]])
    writer.writerow(["Email", data["user"]["email"]])
    writer.writerow(["Full Name", data["user"]["full_name"]])
    writer.writerow(["Role", data["user"]["role"]])
    writer.writerow(["Export Date", data["export_date"]])
    writer.writerow([])
    
    # Сессии
    writer.writerow(["Sessions"])
    writer.writerow(["ID", "Student ID", "Mentor ID", "Scheduled At", "Duration", "Status"])
    for session in data["sessions"]:
        writer.writerow([
            session["id"],
            session["student_id"],
            session["mentor_id"],
            session["scheduled_at"],
            session["duration_minutes"],
            session["status"]
        ])
    writer.writerow([])
    
    # Платежи
    writer.writerow(["Payments"])
    writer.writerow(["ID", "Student ID", "Mentor ID", "Amount", "Currency", "Status", "Created At"])
    for payment in data["payments"]:
        writer.writerow([
            payment["id"],
            payment["student_id"],
            payment["mentor_id"],
            payment["amount"],
            payment["currency"],
            payment["status"],
            payment["created_at"]
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="mentorhub_data_export_{data["user"]["username"]}_{datetime.utcnow().strftime("%Y%m%d")}.csv"'
        }
    )


@router.get("/data/summary", summary="Краткая статистика данных пользователя")
@cached(ttl=300)  # Кэшируем на 5 минут
async def get_data_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Краткая статистика о данных пользователя без детализации.
    """
    
    # Подсчитываем количество записей
    sessions_count = db.execute(
        select(SessionModel).where(
            (SessionModel.student_id == current_user.id) | 
            (SessionModel.mentor_id == current_user.id)
        )
    ).scalars().count()
    
    payments_count = db.execute(
        select(Payment).where(
            (Payment.student_id == current_user.id) | 
            (Payment.mentor_id == current_user.id)
        )
    ).scalars().count()
    
    reviews_count = db.execute(
        select(Review).where(Review.user_id == current_user.id)
    ).scalars().count()
    
    progress_count = db.execute(
        select(Progress).where(Progress.user_id == current_user.id)
    ).scalars().count()
    
    achievements_count = db.execute(
        select(Achievement).where(Achievement.user_id == current_user.id)
    ).scalars().count()
    
    messages_count = db.execute(
        select(Message).where(
            (Message.sender_id == current_user.id) | 
            (Message.recipient_id == current_user.id)
        )
    ).scalars().count()
    
    enrollments_count = db.execute(
        select(CourseEnrollment).where(CourseEnrollment.user_id == current_user.id)
    ).scalars().count()
    
    return {
        "user_id": current_user.id,
        "username": current_user.username,
        "data_summary": {
            "sessions": sessions_count,
            "payments": payments_count,
            "reviews": reviews_count,
            "progress_records": progress_count,
            "achievements": achievements_count,
            "messages": messages_count,
            "course_enrollments": enrollments_count,
        },
        "total_records": (
            sessions_count + payments_count + reviews_count +
            progress_count + achievements_count + messages_count + enrollments_count
        )
    }


def _export_as_pdf(data: dict) -> Response:
    """Экспорт данных в PDF формат"""
    if not REPORTLAB_AVAILABLE:
        raise HTTPException(status_code=503, detail="PDF export not available (reportlab not installed)")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.5*inch, leftMargin=0.5*inch)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a365d'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    story = []
    
    # Заголовок
    story.append(Paragraph("MentorHub Data Export", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Информация о пользователе
    user_info = [
        ["Username", data["user"]["username"]],
        ["Email", data["user"]["email"]],
        ["Full Name", data["user"]["full_name"] or "N/A"],
        ["Role", data["user"]["role"]],
        ["Export Date", data["export_date"]]
    ]
    
    user_table = Table(user_info, colWidths=[2*inch, 4*inch])
    user_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a90d9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f7ff')])
    ]))
    story.append(user_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Статистика
    stats_data = [
        ["Category", "Count"],
        ["Sessions", len(data["sessions"])],
        ["Payments", len(data["payments"])],
        ["Reviews", len(data["reviews"])],
        ["Progress Records", len(data["progress"])],
        ["Achievements", len(data["achievements"])],
        ["Messages", len(data["messages"])],
        ["Course Enrollments", len(data["enrollments"])]
    ]
    
    stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a90d9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f7ff')])
    ]))
    story.append(stats_table)
    
    doc.build(story)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="mentorhub_export_{data["user"]["username"]}_{datetime.utcnow().strftime("%Y%m%d")}.pdf"'
        }
    )


def _export_as_excel(data: dict) -> Response:
    """Экспорт данных в Excel формат"""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=503, detail="Excel export not available (openpyxl not installed)")
    
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left_side=Side(style='thin'),
        right_side=Side(style='thin'),
        top_side=Side(style='thin'),
        bottom_side=Side(style='thin')
    )
    
    # Sheet 1: User Profile
    ws_user = wb.active
    ws_user.title = "User Profile"
    
    user_data = [
        ["Field", "Value"],
        ["Username", data["user"]["username"]],
        ["Email", data["user"]["email"]],
        ["Full Name", data["user"]["full_name"] or "N/A"],
        ["Role", data["user"]["role"]],
        ["Is Active", str(data["user"]["is_active"])],
        ["Is Verified", str(data["user"]["is_verified"])],
        ["Export Date", data["export_date"]]
    ]
    
    for row_idx, row in enumerate(user_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws_user.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            else:
                cell.border = thin_border
    
    # Настройка ширины колонок
    ws_user.column_dimensions['A'].width = 20
    ws_user.column_dimensions['B'].width = 40
    
    # Sheet 2: Sessions
    ws_sessions = wb.create_sheet(title="Sessions")
    sessions_data = [["ID", "Student ID", "Mentor ID", "Scheduled At", "Duration", "Status"]]
    for session in data["sessions"]:
        sessions_data.append([
            session["id"],
            session["student_id"],
            session["mentor_id"],
            session["scheduled_at"],
            session["duration_minutes"],
            session["status"]
        ])
    
    _write_excel_sheet(ws_sessions, sessions_data)
    
    # Sheet 3: Payments
    ws_payments = wb.create_sheet(title="Payments")
    payments_data = [["ID", "Student ID", "Mentor ID", "Amount", "Currency", "Status", "Created At"]]
    for payment in data["payments"]:
        payments_data.append([
            payment["id"],
            payment["student_id"],
            payment["mentor_id"],
            payment["amount"],
            payment["currency"],
            payment["status"],
            payment["created_at"]
        ])
    
    _write_excel_sheet(ws_payments, payments_data)
    
    # Sheet 4: Statistics
    ws_stats = wb.create_sheet(title="Statistics")
    stats_data = [
        ["Category", "Count"],
        ["Sessions", len(data["sessions"])],
        ["Payments", len(data["payments"])],
        ["Reviews", len(data["reviews"])],
        ["Progress Records", len(data["progress"])],
        ["Achievements", len(data["achievements"])],
        ["Messages", len(data["messages"])],
        ["Course Enrollments", len(data["enrollments"])],
        ["Total", sum([
            len(data["sessions"]),
            len(data["payments"]),
            len(data["reviews"]),
            len(data["progress"]),
            len(data["achievements"]),
            len(data["messages"]),
            len(data["enrollments"])
        ])]
    ]
    
    _write_excel_sheet(ws_stats, stats_data)
    
    # Сохраняем
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="mentorhub_export_{data["user"]["username"]}_{datetime.utcnow().strftime("%Y%m%d")}.xlsx"'
        }
    )


def _write_excel_sheet(ws, data):
    """Вспомогательная функция для записи Excel листа"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left_side=Side(style='thin'),
        right_side=Side(style='thin'),
        top_side=Side(style='thin'),
        bottom_side=Side(style='thin')
    )
    
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            else:
                cell.border = thin_border
        
        # Настройка ширины колонок
        for col_idx in range(1, len(data[0]) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
